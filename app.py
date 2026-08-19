"""
تطبيق حجز المرضى — عيادة الحرش الطبية
=========================================
هذا الجزء فقط مخصص للعمل عبر المتصفح (Chrome) — يحتوي فقط على صفحة
الحجز العمومية (/) واستقبال الطلب (/book). كل ما يخص الاستقبال
والأطباء والفواتير أصبح في تطبيق سطح المكتب المنفصل (desktop_app).

يشارك هذا التطبيق نفس قاعدة البيانات (SQLite) مع تطبيق سطح المكتب عبر
مجلد "data" المشترك — انظر config.py.
"""

import os
from datetime import datetime
from pathlib import Path

import psycopg2
from psycopg2.extras import RealDictCursor
from flask import (
    Flask, render_template, request, redirect, url_for, flash, g
)
from openpyxl import Workbook, load_workbook

from config import DATABASE_URL, UPLOADS_DIR, EXPORTS_DIR
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "clinique-el-harrach-demo-secret-change-me")

SPECIALTIES = {
    "radiologie": "الأشعة",
    "cardiologie": "القلب",
    "medecine_generale": "الطب العام",
    "ophtalmologie": "العيون",
}

BOOKING_OPEN_HOUR = 6
BOOKING_CLOSE_HOUR = 24

DEMO_PASSWORD = "demo1234"


def specialty_label(key):
    return SPECIALTIES.get(key, key)


app.jinja_env.globals["specialty_label"] = specialty_label


# ---------------------------------------------------------------------------
# Database helpers (نفس مخطط تطبيق سطح المكتب — قاعدة بيانات مشتركة)
# ---------------------------------------------------------------------------
class DatabaseWrapper:
    def __init__(self, url):
        self.conn = psycopg2.connect(url)

    def execute(self, sql, params=None):
        sql = sql.replace("?", "%s")
        cur = self.conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(sql, params or ())
        return cur

    def executemany(self, sql, params_list):
        sql = sql.replace("?", "%s")
        cur = self.conn.cursor(cursor_factory=RealDictCursor)
        cur.executemany(sql, params_list)
        return cur

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def get_db():
    if "db" not in g:
        if not DATABASE_URL:
            raise RuntimeError(
                "DATABASE_URL غير موجود. "
                "يجب ضبط متغير البيئة DATABASE_URL."
            )

        g.db = DatabaseWrapper(DATABASE_URL)

    return g.db
@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    """إنشاء جداول PostgreSQL وبيانات التجربة عند الحاجة."""

    if not DATABASE_URL:
        raise RuntimeError(
            "DATABASE_URL غير موجود. يجب ضبطه قبل تشغيل التطبيق."
        )

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            role TEXT NOT NULL,
            fullname TEXT NOT NULL,
            specialty TEXT,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            phone TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS patients (
            id SERIAL PRIMARY KEY,
            fullname TEXT NOT NULL,
            phone TEXT,
            age INTEGER,
            sexe TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS exam_categories (
            id SERIAL PRIMARY KEY,
            specialty TEXT NOT NULL,
            name TEXT NOT NULL,
            daily_limit INTEGER
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS exam_subtypes (
            id SERIAL PRIMARY KEY,
            category_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            price DOUBLE PRECISION NOT NULL,
            FOREIGN KEY(category_id)
                REFERENCES exam_categories(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS appointments (
            id SERIAL PRIMARY KEY,
            patient_id INTEGER NOT NULL,
            doctor_id INTEGER,
            specialty TEXT NOT NULL,
            exam TEXT NOT NULL,
            subtype TEXT,
            price DOUBLE PRECISION DEFAULT 0,
            paid DOUBLE PRECISION DEFAULT 0,
            remaining DOUBLE PRECISION DEFAULT 0,
            status TEXT DEFAULT 'en_attente',
            prescription TEXT,
            invoice_number TEXT,
            invoice_date TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(patient_id)
                REFERENCES patients(id),
            FOREIGN KEY(doctor_id)
                REFERENCES users(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS prescriptions (
            id SERIAL PRIMARY KEY,
            patient_id INTEGER NOT NULL,
            doctor_id INTEGER NOT NULL,
            appointment_id INTEGER,
            date TEXT NOT NULL,
            medicaments TEXT,
            notes TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id SERIAL PRIMARY KEY,
            invoice_number TEXT UNIQUE,
            appointment_id INTEGER,
            total_ht DOUBLE PRECISION,
            tva DOUBLE PRECISION,
            total_ttc DOUBLE PRECISION,
            created_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS clinic_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            name TEXT,
            address TEXT,
            phone TEXT,
            city TEXT,
            matricule_fiscal TEXT,
            matricule_stat TEXT,
            article_imposition TEXT,
            compte_bancaire TEXT
        )
    """)

    conn.commit()

    cur.execute("SELECT COUNT(*) FROM users")
    count = cur.fetchone()[0]

    if count == 0:
        seed_demo_data(cur)
        conn.commit()

    cur.close()
    conn.close()
def seed_demo_data(cur):
    cur.execute(
        """
        INSERT INTO clinic_settings
        (id, name, address, phone, city, matricule_fiscal,
         matricule_stat, article_imposition, compte_bancaire)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            1,
            "Clinique El Harrach Médicale",
            "QUARTIER TIMAKRET CO4 ET CO5 - METLILI CHAAMBA",
            "020.71.86.82 / 020.71.87.29",
            "METLILI CHAAMBA",
            "184470500605171",
            "198447050060526",
            "47050014916",
            "00300299000043330082",
        ),
    )

    staff = [
        (
            "reception",
            "الاستقبال",
            None,
            "demo.reception@clinique.com",
            DEMO_PASSWORD,
            "",
        ),
        (
            "medecin",
            "Dr LAHRECHE HOUARI",
            "radiologie",
            "demo.radio@clinique.com",
            DEMO_PASSWORD,
            "",
        ),
        (
            "medecin",
            "Dr AMINE BELABBES",
            "cardiologie",
            "demo.cardio@clinique.com",
            DEMO_PASSWORD,
            "",
        ),
        (
            "medecin",
            "Dr SARAH MERABET",
            "medecine_generale",
            "demo.generale@clinique.com",
            DEMO_PASSWORD,
            "",
        ),
        (
            "medecin",
            "Dr YOUCEF ZEROUKI",
            "ophtalmologie",
            "demo.ophta@clinique.com",
            DEMO_PASSWORD,
            "",
        ),
    ]

    cur.executemany(
        """
        INSERT INTO users
        (role, fullname, specialty, email, password, phone)
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        staff,
    )

    catalog = [
        (
            "radiologie",
            "Scanner",
            2,
            [
                ("Brain", 7000),
                ("Abdomen", 6500),
                ("Thorax", 6000),
            ],
        ),
        (
            "radiologie",
            "Échographie",
            2,
            [
                ("Échographie Abdominal", 2000),
                ("Échographie Pelvic", 3500),
            ],
        ),
        (
            "radiologie",
            "Mammographie",
            1,
            [
                ("Mammographie", 5000),
            ],
        ),
        (
            "radiologie",
            "Radiographie",
            None,
            [
                ("Standard", 2000),
                ("Spine", 3000),
            ],
        ),
        (
            "radiologie",
            "Panoramique",
            None,
            [
                ("Panoramique Dentaire", 2500),
            ],
        ),
        (
            "cardiologie",
            "Électrocardiogramme",
            None,
            [
                ("ECG standard", 1500),
            ],
        ),
        (
            "cardiologie",
            "Échocardiographie",
            5,
            [
                ("Échocardiographie transthoracique", 4000),
                ("Échocardiographie de stress", 6000),
            ],
        ),
        (
            "cardiologie",
            "Test d'effort",
            3,
            [
                ("Épreuve d'effort", 5000),
            ],
        ),
        (
            "cardiologie",
            "Holter ECG",
            2,
            [
                ("Holter 24h", 4500),
            ],
        ),
        (
            "cardiologie",
            "Consultation cardiologique",
            None,
            [
                ("Consultation", 2000),
            ],
        ),
        (
            "medecine_generale",
            "Consultation générale",
            None,
            [
                ("Consultation", 1500),
            ],
        ),
        (
            "medecine_generale",
            "Consultation de contrôle",
            None,
            [
                ("Contrôle", 1000),
            ],
        ),
        (
            "medecine_generale",
            "Certificat médical",
            None,
            [
                ("Certificat", 800),
            ],
        ),
        (
            "medecine_generale",
            "Vaccination",
            None,
            [
                ("Injection / Vaccin", 500),
            ],
        ),
        (
            "ophtalmologie",
            "Consultation ophtalmologique",
            None,
            [
                ("Consultation", 2000),
            ],
        ),
        (
            "ophtalmologie",
            "Examen de la vue",
            None,
            [
                ("Réfraction", 1500),
            ],
        ),
        (
            "ophtalmologie",
            "Fond d'œil",
            10,
            [
                ("Fond d'œil", 2500),
            ],
        ),
        (
            "ophtalmologie",
            "Champ visuel",
            5,
            [
                ("Champ visuel", 3000),
            ],
        ),
    ]

    for specialty, name, limit, subtypes in catalog:
        cur.execute(
            """
            INSERT INTO exam_categories
            (specialty, name, daily_limit)
            VALUES (%s, %s, %s)
            RETURNING id
            """,
            (specialty, name, limit),
        )

        category_id = cur.fetchone()[0]

        for sub_name, price in subtypes:
            cur.execute(
                """
                INSERT INTO exam_subtypes
                (category_id, name, price)
                VALUES (%s, %s, %s)
                """,
                (category_id, sub_name, price),
            )
# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def is_open():
    now = datetime.now()
    return BOOKING_OPEN_HOUR <= now.hour < BOOKING_CLOSE_HOUR


def exams_by_specialty_dict():
    db = get_db()
    rows = db.execute("SELECT specialty, name FROM exam_categories ORDER BY id").fetchall()
    result = {key: [] for key in SPECIALTIES}
    for r in rows:
        result.setdefault(r["specialty"], []).append(r["name"])
    return result


def get_category(specialty, exam_name):
    db = get_db()
    return db.execute(
        "SELECT * FROM exam_categories WHERE specialty=? AND name=?", (specialty, exam_name)
    ).fetchone()


def get_default_price(category_id):
    db = get_db()
    row = db.execute(
        "SELECT price FROM exam_subtypes WHERE category_id=? ORDER BY id LIMIT 1", (category_id,)
    ).fetchone()
    return row["price"] if row else 0


def get_exam_usage_today(specialty):
    db = get_db()
    today = datetime.now().date()

    rows = db.execute(
        """
        SELECT exam, COUNT(*) AS c
        FROM appointments
        WHERE specialty = ?
          AND created_at::date = ?
          AND status != 'annule'
        GROUP BY exam
        """,
        (specialty, today),
    ).fetchall()

    return {r["exam"]: r["c"] for r in rows}

def get_daily_folder(base):
    today = datetime.now()
    folder = Path(base) / str(today.year) / f"{today.month:02d}" / f"{today.day:02d}"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def clean_filename(name):
    return "".join(c for c in name if c.isalnum() or c in (" ", "_")).strip().replace(" ", "_")


def find_or_create_patient(fullname, phone, age):
    db = get_db()

    existing = db.execute(
        "SELECT * FROM patients WHERE phone=?",
        (phone,)
    ).fetchone()

    if existing:
        return existing["id"]

    cur = db.execute(
        """
        INSERT INTO patients
        (fullname, phone, age, created_at)
        VALUES (?, ?, ?, ?)
        RETURNING id
        """,
        (
            fullname,
            phone,
            age,
            datetime.now().isoformat(sep=" ")
        ),
    )

    patient_id = cur.fetchone()["id"]
    db.commit()

    return patient_id

def save_to_excel(fullname, phone, specialty, exam, price, paid, remaining, created_at):
    folder = get_daily_folder(EXPORTS_DIR)
    file_path = folder / "patients.xlsx"
    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Specialty", "Exam", "Price", "Paid", "Remaining", "Date"])
    ws.append([fullname, phone, specialty_label(specialty), exam, price, paid, remaining, created_at])
    wb.save(file_path)


# ---------------------------------------------------------------------------
# Public booking (الصفحتان الوحيدتان في هذا التطبيق)
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template(
        "index.html", open=is_open(),
        specialties=SPECIALTIES, exams_by_specialty=exams_by_specialty_dict(),
    )

@app.route("/debug-db")
def debug_db():
    db = get_db()
    info = db.execute("SELECT current_database() AS dbname, inet_server_addr()::text AS ip").fetchone()
    count_p = db.execute("SELECT COUNT(*) AS c FROM patients").fetchone()
    count_a = db.execute("SELECT COUNT(*) AS c FROM appointments").fetchone()
    last = db.execute("SELECT id, created_at FROM appointments ORDER BY id DESC LIMIT 1").fetchone()
    return {
        "database": info["dbname"],
        "server_ip": info["ip"],
        "patients_count": count_p["c"],
        "appointments_count": count_a["c"],
        "last_appointment": dict(last) if last else None,
    }
@app.route("/book", methods=["POST"])
def book():
    if not is_open():
        flash("❌ الحجز مغلق حالياً", "error")
        return redirect(url_for("index"))

    db = get_db()
    fullname = request.form["fullname"]
    phone = request.form["phone"]
    specialty = request.form["specialty"]
    exam = request.form["exam"]
    age = request.form.get("age", 0) or 0

    if specialty not in SPECIALTIES:
        flash("❌ تخصص غير صالح", "error")
        return redirect(url_for("index"))

    category = get_category(specialty, exam)
    if not category:
        flash("❌ فحص غير صالح", "error")
        return redirect(url_for("index"))

    if category["daily_limit"] is not None:
        usage = get_exam_usage_today(specialty).get(exam, 0)
        if usage >= category["daily_limit"]:
            flash(f"❌ تم الوصول للحد الأقصى لهذا الفحص اليوم ({exam})", "error")
            return redirect(url_for("index"))

    price = get_default_price(category["id"])
    now = datetime.now().isoformat(sep=" ")
    patient_id = find_or_create_patient(fullname, phone, age)

    filename = None
    file = request.files.get("prescription")
    if file and file.filename != "":
        folder = get_daily_folder(UPLOADS_DIR)
        safe_name = clean_filename(fullname)
        ext = file.filename.rsplit(".", 1)[-1].lower()
        filepath = folder / f"{safe_name}_{int(datetime.now().timestamp())}.{ext}"
        file.save(str(filepath))
        filename = str(filepath.relative_to(UPLOADS_DIR))

    db.execute("""
        INSERT INTO appointments (patient_id, specialty, exam, price, paid, remaining,
                                    status, prescription, created_at)
        VALUES (?, ?, ?, ?, 0, ?, 'en_attente', ?, ?)
    """, (patient_id, specialty, exam, price, price, filename, now))
    db.commit()

    save_to_excel(fullname, phone, specialty, exam, price, 0, price, now)
    flash("✅ تم إرسال طلبك بنجاح، سيتم استدعاؤكم من طرف الاستقبال")
    return redirect(url_for("index"))


# تهيئة قاعدة البيانات عند تشغيل التطبيق
# يعمل سواءً شُغّل التطبيق مباشرة أو بواسطة Gunicorn على Render
init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)